import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference

from evm_logger import get_logger
from database import get_vote_totals, get_all_events

log = get_logger("smart_evm.export")


def _get_export_dir() -> str:
    try:
        from kivy.app import App
        app = App.get_running_app()
        if app is not None:
            path = os.path.join(app.user_data_dir, "exports")
            os.makedirs(path, exist_ok=True)
            return path
    except Exception:
        pass
    path = os.path.join(os.path.expanduser("~"), "SmartEVM_exports")
    os.makedirs(path, exist_ok=True)
    return path


def _header_style(cell, bg: str = "1F2937", fg: str = "FFFFFF") -> None:
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def export(candidate_map: dict, out_path: str = None) -> str:
    if out_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(_get_export_dir(), f"results_{ts}.xlsx")

    os.makedirs(os.path.dirname(out_path) if os.path.dirname(out_path) else ".", exist_ok=True)
    totals = get_vote_totals()
    events = get_all_events()

    wb = openpyxl.Workbook()

    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 6
    ws_sum.column_dimensions["B"].width = 22
    ws_sum.column_dimensions["C"].width = 14

    title_cell = ws_sum["A1"]
    title_cell.value = "SMART EVM — Vote Summary"
    title_cell.font = Font(bold=True, size=14, color="1A56DB")
    ws_sum.merge_cells("A1:C1")
    title_cell.alignment = Alignment(horizontal="center")

    ws_sum["A2"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_sum["A2"].font = Font(italic=True, color="6B7280", size=9)
    ws_sum.merge_cells("A2:C2")

    for col, h in enumerate(["ID", "Candidate", "Votes"], start=1):
        _header_style(ws_sum.cell(row=4, column=col, value=h))

    total_votes = 0
    for row_idx, (cid, name) in enumerate(candidate_map.items(), start=5):
        votes = totals.get(cid, 0)
        total_votes += votes
        ws_sum.cell(row=row_idx, column=1, value=cid)
        ws_sum.cell(row=row_idx, column=2, value=name)
        ws_sum.cell(row=row_idx, column=3, value=votes)

    footer_row = 5 + len(candidate_map)
    ws_sum.cell(row=footer_row, column=2, value="TOTAL").font = Font(bold=True)
    ws_sum.cell(row=footer_row, column=3, value=total_votes).font = Font(bold=True)

    data_ref = Reference(ws_sum, min_col=3, min_row=4, max_row=4 + len(candidate_map))
    cats_ref = Reference(ws_sum, min_col=2, min_row=5, max_row=4 + len(candidate_map))

    bar = BarChart()
    bar.type = "col"
    bar.title = "Votes per Candidate"
    bar.y_axis.title = "Votes"
    bar.x_axis.title = "Candidate"
    bar.style = 10
    bar.width = 18
    bar.height = 12
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    ws_sum.add_chart(bar, "E4")

    pie = PieChart()
    pie.title = "Vote Distribution"
    pie.style = 10
    pie.width = 14
    pie.height = 12
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    ws_sum.add_chart(pie, "E22")

    ws_log = wb.create_sheet("Event Log")
    ws_log.column_dimensions["A"].width = 6
    ws_log.column_dimensions["B"].width = 22
    ws_log.column_dimensions["C"].width = 10
    ws_log.column_dimensions["D"].width = 22
    ws_log.column_dimensions["E"].width = 14

    for col, h in enumerate(["ID", "Timestamp", "Candidate ID", "Candidate Name", "Event Type"], start=1):
        _header_style(ws_log.cell(row=1, column=col, value=h))

    for r_idx, row in enumerate(reversed(events), start=2):
        ws_log.cell(row=r_idx, column=1, value=row["id"])
        ws_log.cell(row=r_idx, column=2, value=row["timestamp"])
        ws_log.cell(row=r_idx, column=3, value=row["candidate_id"])
        ws_log.cell(row=r_idx, column=4, value=row["candidate_name"])
        ws_log.cell(row=r_idx, column=5, value=row["event_type"])

    wb.save(out_path)
    abs_path = os.path.abspath(out_path)
    log.info("Excel exported → %s", abs_path)
    return abs_path
